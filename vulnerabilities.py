#!/usr/bin/python
#--------------------------------------------------------------------------------------------------
# vulnerabilities.py: Python module used by CveAdvisory.py
#--------------------------------------------------------------------------------------------------
#import sqlite3 as lite
#from vfeed import vFeed, vFeedUpdate
import MySQLdb as mysql
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Color
from openpyxl.comments import Comment
#from openpyxl.charts import BarChart, Reference, Series
#from openpyxl.drawing import Drawing, Image
import calendar
import re
import urllib
import sys
import socket
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.MIMEBase import MIMEBase
from email import Encoders
import urllib2
import AdvisoryConfig as cfg
import os
import subprocess
from BeautifulSoup import BeautifulSoup
import xml.etree.ElementTree as ET

#--------------------------------------------------------------------------------------------------
# Get the NVD XML file from the remote server
#--------------------------------------------------------------------------------------------------
def getCveDataFromNVD(xmlfile):
    url 	= cfg.nvdCveXmlBase + xmlfile
    filename 	= cfg.repDir        + xmlfile
    
    u = urllib2.urlopen(url)
    f = open(filename, 'wb')
    meta = u.info()
    filesize = int(meta.getheaders("Content-Length")[0])
    filesize_dl = 0
    block_sz = 8192
    while True:
	sys.stdout.flush()
	buffer = u.read(block_sz)
	if not buffer:
	    break
	filesize_dl += len(buffer)
	f.write(buffer)
	status = r"%10d [%3.0f %%]" % (filesize_dl, filesize_dl * 100. / filesize)
	status = status + chr(8)*(len(status)+1)
	sys.stdout.write("\r[progress %3.0f %%] receiving %d out of %s Bytes of %s " % (filesize_dl * 100. / filesize, filesize_dl,filesize,filename))
	sys.stdout.flush()
    f.close()
    print "-- Download completed."


#--------------------------------------------------------------------------------------------------
# Parse NVD XML file and generate a list to be used for Excel export
#--------------------------------------------------------------------------------------------------
def processNvdXML(xmlfile):

    tree = ET.parse(cfg.repDir + xmlfile)
    nsVuln='{http://scap.nist.gov/schema/vulnerability/0.4}'
    nsCvss='{http://scap.nist.gov/schema/cvss-v2/0.2}'
    nsCpeLang='{http://cpe.mitre.org/language/2.0}'
    VulList = []    
    
    root = tree.getroot()
    for cve in root:
	cveId 	= cve.attrib['id']
	summary = cve.find('{0}summary'.format(nsVuln)).text
	if len(summary) >= 255:
	    summary = summary[:251] + "..."
	if summary[:12] == "** REJECT **":
	    continue
	print cveId;
	pubDate	= cve.find('{0}published-datetime'.format(nsVuln)).text[:10]
	pubYear	= pubDate[:4]
	pubMonth= pubDate[5:7]
	modDate	= cve.find('{0}last-modified-datetime'.format(nsVuln)).text[:10]
	try:
	    score  	= cve.find('{0}cvss/{1}base_metrics/{1}score'.format(nsVuln, nsCvss)).text
	except AttributeError:
	    score	= "-1.0"
	    
	if (float(score) < 0.0):
	    severity = cfg.tbd
	elif (float(score) < 4.0):
	    severity = 'Low'
	elif (float(score) < 7.0):
	    severity = 'Medium'
	else:
	    severity = 'High'
	    
	vulnCfgs=""
	for vc in cve.findall('{0}vulnerable-software-list/{0}product'.format(nsVuln)):
	    if vulnCfgs != "":
		vulnCfgs += "\n"
	    vulnCfgs += vc.text[7:]
	
	try:
	    auth	= cve.find('{0}cvss/{1}base_metrics/{1}authentication'.format(nsVuln,nsCvss)).text
	except AttributeError:
	    auth	= cfg.tbd
	    
	try:
	    aComplexity	= cve.find('{0}cvss/{1}base_metrics/{1}access-complexity'.format(nsVuln,nsCvss)).text
	except AttributeError:
	    aComplexity	= cfg.tbd
	    
	try:
	    cImpact	= cve.find('{0}cvss/{1}base_metrics/{1}confidentiality-impact'.format(nsVuln,nsCvss)).text
	except AttributeError:
	    cImpact	= cfg.tbd

	try:
	    iImpact	= cve.find('{0}cvss/{1}base_metrics/{1}integrity-impact'.format(nsVuln,nsCvss)).text
	except AttributeError:
	    iImpact	= cfg.tbd

	try:
	    aImpact	= cve.find('{0}cvss/{1}base_metrics/{1}availability-impact'.format(nsVuln,nsCvss)).text
	except AttributeError:
	    aImpact	= cfg.tbd

	try:
	    accessV	= cve.find('{0}cvss/{1}base_metrics/{1}access-vector'.format(nsVuln,nsCvss)).text
	except AttributeError:
	    accessV	= cfg.tbd

	try:
	    cweId	= cve.find('{0}cwe'.format(nsVuln)).attrib['id']
	except AttributeError:
	    cweId	= cfg.tbd

	refs=""
	for ref in cve.findall('{0}references'.format(nsVuln)):
	    if refs != "":
		refs += "\n"    
	    refs += ref.find('{0}source'.format(nsVuln)).text + ":" + ref.find('{0}reference'.format(nsVuln)).text

	# Push all CVE data to results list
	result = (cveId,pubDate,modDate,summary,score,severity,vulnCfgs,auth,aComplexity,cImpact,iImpact,aImpact,accessV,cweId,refs)
	VulList.append(result)
    return VulList    


#--------------------------------------------------------------------------------------------------
# Update local instance of this year's CVE database
#--------------------------------------------------------------------------------------------------
def updateLocalCveDB(vulns, tblName):
    
    print "-- Connecting to the database " + cfg.dbName
    db = mysql.connect(cfg.dbHost, cfg.dbUser, cfg.dbPwd, cfg.dbName)
    cur = db.cursor()
    
    # Truncate the table 
    print "-- Truncating the CVE table, prior to (re)inserting data"
    cur.execute("TRUNCATE TABLE " + tblName)

    print "-- Inserting " , len(vulns) , " CVE records"
    for vuln in vulns:
	cols=""
	for col in cfg.nvdFields:
	    if cols != "":
		cols += ","
	    cols+=col
	sql = "INSERT INTO %s(%s) VALUES(%%s,%%s,%%s,%%s,%%s,%%s,%%s,%%s,%%s,%%s,%%s,%%s,%%s,%%s,%%s)" %(tblName, cols)
	#print "--- Inserting " + vuln[0]
	try:
		cur.execute(sql, vuln)
		db.commit()
		#print "Last Executed: ", cur._last_executed 
	except Exception as ex:
		template = "Exception of type {0} occured. Arguments:\n{1!r}"
		message = template.format(type(ex).__name__, ex.args)
		print message	    
    print "-- Done."

    # Disconnect from DB server
    db.close()    
    
    

#--------------------------------------------------------------------------------------------------
# Get advisory data from the local CVE database
#--------------------------------------------------------------------------------------------------
def getCVEsByAsset(assetList, startDate, endDate):
    print "-- Connecting to the database " + cfg.dbName
    db = mysql.connect(cfg.dbHost, cfg.dbUser, cfg.dbPwd, cfg.dbName)
    cur = db.cursor()
    vulList = []
    
    for asset in assetList:
        assetSearch = "%" + asset + "%"
	print "-- Getting CVEs for assets matching: " + asset
	
	# Query the DB
	Query = '''SELECT 
			    YEAR(Date_Published) 	as 'c0', 
			    MONTH(Date_Published) 	as 'c1',
			    Date_Published	 	as 'c2',
			    CVE_ID 			as 'c3',
			    Title		 	as 'c4', 
			    Last_Modified	 	as 'c5',
			    CVSS_Score 			as 'c6',
			    Severity 			as 'c7',
			    Affected_Software 		as 'c8',
			    External_References 	as 'c9',
			    Authentication_Required	as 'c10',
			    Access_Complexity		as 'c11',
			    Confidentiality_Impact 	as 'c12',
			    Integrity_Impact 		as 'c13',
			    Availability_Impact 	as 'c14',
			    Access_Vector 		as 'c15',
			    CWE_ID 			as 'c16'
		    FROM cve
		    WHERE 		Date_Published >= '%s'
			    AND 	Date_Published <= '%s'
			    AND 	Affected_Software LIKE '%s'
		    ORDER by Date_Published DESC, CVE_ID DESC;	
                ''' % (startDate, endDate, assetSearch)

        cur.execute(Query)
        result = cur.fetchall()
        for row in result:
            # Get data from SQL query
	    year        = row[0]
	    month     	= row[1]
	    publishDate	= row[2]
	    cveId     	= row[3]
	    title      	= row[4]
	    modDate    	= row[5]
	    cvssScore  	= row[6]
	    severity   	= row[7]
	    affectedSW	= row[8]
	    refs      	= row[9]
	    authReq    	= row[10]
	    aComplexity	= row[11]
	    cImpact    	= row[12]
	    iImpact    	= row[13]
	    aImpact    	= row[14]
	    aVector    	= row[15]
	    cweId	= row[16]
	    cveURI    	= cfg.cveUrlBase + cveId
	    cweURI    	= ""
	    if cweId != cfg.tbd:
		cweURI = cfg.cweUrlBase + cweId[4:]

            # Push all CVE data to results list
	    #print " - " + cveId
            #result = (asset,year,month,publishDate,cveId,title,modDate,cveURI,severity,affectedSW,refs,authReq,aComplexity,cImpact,iImpact,aImpact,aVector,cvssScore,cweURI)
            result = (year,month,publishDate,cveId,asset,title,modDate,cveURI,severity,affectedSW,refs,authReq,aComplexity,cImpact,iImpact,aImpact,aVector,cvssScore,cweURI)
            vulList.append(result)
    print "-- Done. Total records: ", len(vulList)
    return vulList


#--------------------------------------------------------------------------------------------------
# Create an Excel report from NVD data
#--------------------------------------------------------------------------------------------------
def createCVEsByAssetReport(domain, vulnList, year, month, filename):

    # Load Excel template
    destFilename = "reports/" + filename
    wb = load_workbook(filename = cfg.excelTemplate, use_iterators=False, keep_vba=False, guess_types=False, data_only=False)
    ws1 = wb.get_sheet_by_name("Advisories")
    ws2 = wb.get_sheet_by_name("Monthly Statistics")    


    # Write data content    
    print "-- Adding data to spreadsheet"
    rCounter = 7
    for vuln in vulnList:
        cCounter = 1        
        for item in vuln:
	    ws1.cell(row = rCounter, column = cCounter).value = item
	    if "URL" in cfg.rptFields[cCounter-1] and item[:4] == "http":
		ws1.cell(row = rCounter, column = cCounter).value = '=hyperlink("' + item + '","' + item + '")'
	    cCounter += 1
        rCounter += 1
    ws2["B5"].value = int(month)-1 

    # Add header images 
    #img1 = Image('logo1-small.gif')
    #img2 = Image('logo2-small.png')
    #img1.drawing.top = 2
    #img1.drawing.left = 0
    #img2.drawing.top = 2
    #img2.drawing.left = 600
    #ws1.add_image(img1)
    #ws1.add_image(img2)    
    #ws2.add_image(img1)
    #ws2.add_image(img2)
    
    # Save the Excel workbook
    wb.save(filename = destFilename)
    print "-- Report file name generated: " + destFilename


#--------------------------------------------------------------------------------------------------
# Create an Excel report from NVD data
#--------------------------------------------------------------------------------------------------
def createSimpleCVEReport(rptYear, filename, dbTable):
    
    destFilename = "reports/" + filename
    wb = Workbook()
    ws = wb.get_active_sheet()
    ws.title ="%s CVEs" %(rptYear)
    
    print "-- Connecting to the database " + cfg.dbName
    db = mysql.connect(cfg.dbHost, cfg.dbUser, cfg.dbPwd, cfg.dbName)
    cur = db.cursor()
    
    print "-- Getting all CVEs from DB"
    Query = '''SELECT 
			YEAR(Date_Published) 	as 'c0', 
			MONTH(Date_Published) 	as 'c1',
			Date_Published	 	as 'c2',
			CVE_ID 			as 'c3',
			Title		 	as 'c4', 
			Last_Modified	 	as 'c5',
			CVSS_Score 		as 'c6',
			Severity 		as 'c7',
			Affected_Software 	as 'c8',
			External_References 	as 'c9',
			Authentication_Required	as 'c10',
			Access_Complexity	as 'c11',
			Confidentiality_Impact 	as 'c12',
			Integrity_Impact 	as 'c13',
			Availability_Impact 	as 'c14',
			Access_Vector 		as 'c15',
			CWE_ID 			as 'c16'
		FROM %s
		ORDER by Date_Published DESC;	
	    ''' %(dbTable)
    cur.execute(Query)
    result = cur.fetchall()
    vulnList = []
    for row in result:
	# Get data from SQL query
	year        	= row[0]
	month     	= row[1]
	publishDate	= row[2]
	cveId     	= row[3]
	title      	= row[4]
	modDate    	= row[5]
	cvssScore  	= row[6]
	severity   	= row[7]
	affectedSW	= row[8]
	refs      	= row[9]
	authReq    	= row[10]
	aComplexity    	= row[11]
	cImpact    	= row[12]
	iImpact    	= row[13]
	aImpact    	= row[14]
	aVector    	= row[15]
	cweId		= row[16]
	cveURI    	= cfg.cveUrlBase + cveId
	cweURI    	= ""
	if cweId != cfg.tbd:
	    cweURI = cfg.cweUrlBase + cweId[4:]


	# Push all CVE data to results list
	#print " - " + cveId
	result = (year,month,publishDate,cveId,title,modDate,cveURI,severity,affectedSW,refs,authReq,aComplexity,cImpact,iImpact,aImpact,aVector,cvssScore,cweURI)
	vulnList.append(result)

 
    # Write heading row
    print "-- Adding column headers to spreadsheet"
    cCounter = 1
    for field in cfg.rptFieldsYTD:
	comment = Comment(cfg.colComments[cCounter], "(Automated)")
	ws.cell(row = 1, column = cCounter).value = field
	ws.cell(row = 1, column = cCounter).comment = comment
        cCounter += 1

    # Write sheet content    
    print "-- Adding data to spreadsheet"
    rCounter = 2
    for vuln in vulnList:
        cCounter = 1        
        for item in vuln:
	    ws.cell(row = rCounter, column = cCounter).value = unicode(str(item),errors='ignore')

	    if "URL" in cfg.rptFields[cCounter] and item[:4] == "http":
		ws.cell(row = rCounter, column = cCounter).value = '=hyperlink("' + item + '","' + item + '")'
            cCounter += 1
        rCounter += 1
   
    wb.save(filename = destFilename)
    print "-- Report file name generated: " + destFilename


#--------------------------------------------------------------------------------------------------
# Email the Excel report file
#--------------------------------------------------------------------------------------------------
def emailReport(ReportFile, CumulativeReportFile):

    Period 		= ReportFile.split(' ')[0]
    ReportFilePath	= cfg.repDir + ReportFile
    CumulReportFilePath	= cfg.repDir + CumulativeReportFile
    
    print "-- Composing the advisory email"
    
    instructions="""<html>
<head>
    <style>
	tr, td { background-color: #FFFFFF;border: 1px solid #999999;font-family:arial;font-size: 9pt;margin:3px 0px 3px 0px;padding: 2px;vertical-align: top;max-width: 1200px;color: #222222; }
	table{ font-size: 9pt;border-collapse: collapse;color: rgb(128, 128, 128);font-family: Arial,sans-serif;font-style: normal;font-weight: 400;height: 17px; }
	th{ background-attachment: scroll;background-clip: border-box;background-color: rgb(247, 150, 70);background-origin: padding-box;background-position: 0% 0%;background-repeat: repeat;background-size: auto auto;border-bottom-color: rgb(250, 192, 144);border-bottom-style: solid;border-bottom-width: 1px;border-collapse: collapse;border-left-color: rgb(255, 255, 255);border-left-style: none;border-left-width: 0px;border-right-color: rgb(255, 255, 255);border-right-style: none;border-right-width: 0px;border-top-color: rgb(250, 192, 144);border-top-style: solid;border-top-width: 0px;color: rgb(255, 255, 255);font-family: Arial;font-size: 9pt;font-style: normal;font-weight: 700;padding-bottom: 0px;padding-left: 0px;padding-right: 0px;padding-top: 0px;text-decoration: none;white-space: normal; }
	p, div, li{ color: #222222;font-family:arial;font-size: 9pt;text-align: left;vertical-align: top; }
	br,ol,ul,li{ mso-data-placement:same-cell; }
	h3 { color: rgb(100,100,100); font-size: 10pt; }
	a, a:link, a:visited { color: rgb(0, 100, 200); text-decoration: none; }
    </style>
</head>
<body>
<p>Greetings all,</p>

<p>Attached are the <b>Vulnerability Alert Notifications for %s</b>.<br></p>

<h3>Vulnerability Alerts Files</h3>

<p>In the attached files, please review each CVE advisory applicable to the platforms / applications within your area of responsibility,
assess the associated risk, and take timely and appropriate action for your environment.</p>

<p><i>(See attached file: %s)</i><br/><i>(See attached file: %s)</i></p>

<h3>NB:</h3>
<p>
    <ol>
    <li>The first attached file contains CVEs created/updated in the last 60 days. In this file, only the CVEs for matching assets are included, as per the <i>Platform</i> column.</li>
    <li>The second attached file contains all CVEs created/updated this year.</li>
    <li>The <em>References</em> column contains keywords that are described here: http://cve.mitre.org/data/refs/</li>
    <li>For descriptions on columns headers and values, please view this document: http://www.first.org/cvss/cvss-guide</li>
    <li>Searching CVE archives (previous years) can be performed here: http://cve.mitre.org/cve/cve.html or here: http://www.cvedetails.com/product-search.php</li>
    <li>We recommend that the recipients also register with their vendor's advisory sources (email, RSS, Twitter...). A short list is provided below. Do let us know how this list can be improved.</li>
    </ol></p>

<h3>Vendor Security Advisory Notification Services</h3>
<table>
    <tr><td width="160">Adobe Products</td><td width="550">http://helpx.adobe.com/security.html</td></tr>
    <tr><td>Cisco Products</td>        <td>http://tools.cisco.com/security/center/publicationListing</td></tr>
    <tr><td>Fortinet Products</td>     <td>https://support.fortinet.com/login/UserLogin.aspx</td></tr>
    <tr><td>Juniper Products</td>      <td>https://kb.juniper.net/InfoCenter/index?page=subscriptions</td></tr>
    <tr><td>Microsoft MyBulletins</td> <td>http://mybulletins.technet.microsoft.com</td></tr>
    <tr><td>Microsoft Products</td>    <td>https://technet.microsoft.com/en-us/security/bulletin</td></tr>
    <tr><td>Symantec Products</td>     <td>http://www.symantec.com/security_response/securityupdates/list.jsp?fid=security_advisory</td></tr>
    <tr><td>Websense MyWebsense</td>   <td>https://www.websense.com/content/Registration.aspx?task=signin</td></tr>
</table>
</body>
</html>""" % (Period, ReportFile, CumulativeReportFile)
    
    # Send file via email
    print "-- Sending the advisory email to " + cfg.dest
    
    # Create the enclosing (outer) message
    subject="ACTION REQUIRED || Vulnerability Advisory || " + str(Period) + " || Version 1"
    msg = MIMEMultipart()
    msg['From'] = cfg.orig
    msg['To'] = cfg.dest
    msg['Subject'] = subject
    
    # Create a text/plain message with a text file attached
    msg.attach(MIMEText(instructions, 'html'))
    
    # Add first report file
    att1 = MIMEBase('application', "octet-stream")    
    att1.set_payload( open(ReportFilePath,"rb").read() )
    Encoders.encode_base64(att1)
    att1.add_header('Content-Disposition', 'attachment; filename="%s"' % ReportFile)
    msg.attach(att1)
    
    # Add 2nd report file
    att2 = MIMEBase('application', "octet-stream")    
    att2.set_payload( open(CumulReportFilePath,"rb").read() )
    Encoders.encode_base64(att2)
    att2.add_header('Content-Disposition', 'attachment; filename="%s"' % CumulativeReportFile)
    msg.attach(att2)
    
    # Send the message via our own SMTP server, but don't include the envelope header.
    s = smtplib.SMTP(cfg.mx)
    s.sendmail(cfg.orig, cfg.dest, msg.as_string())
    s.quit()
    print "-- Done."
